import os
import pdfquery
import openpyxl
from flask import Flask, request, render_template, send_file, redirect, url_for
import shutil
import threading
import webbrowser

app = Flask(__name__)

def empty_folder(folder_path):
    """Esvazia a pasta especificada."""
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        if os.path.isfile(file_path):
            os.remove(file_path)
    print(f"Pasta '{folder_path}' esvaziada.")

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        # --- 1. Captura dos arquivos ---
        pdf_empenho = request.files.get('pdf_empenho')
        boletos = request.files.getlist('pdf_boletos')

        if not pdf_empenho:
            return "Por favor, envie o PDF da nota de empenho.", 400

        if not boletos:
            return "Por favor, envie ao menos um PDF de boleto.", 400

        # --- 2. Pastas de trabalho ---
        directory_path = os.path.dirname(os.path.abspath(__file__))
        pdfs_directory = os.path.join(directory_path, 'pdfs')
        drafts_directory = os.path.join(directory_path, 'RascunhosGerados')
        os.makedirs(pdfs_directory, exist_ok=True)
        os.makedirs(drafts_directory, exist_ok=True)

        excel_filename = "Modelo.xlsx"
        excel_path = os.path.join(directory_path, excel_filename)
        temp_excel = os.path.join(pdfs_directory, "Temp_Consolidado.xlsx")
        shutil.copy(excel_path, temp_excel)
        wb = openpyxl.load_workbook(temp_excel)
        ws = wb.active

        # --- 3. Extrair dados da nota de empenho ---
        pdf_empenho_path = os.path.join(pdfs_directory, pdf_empenho.filename)
        pdf_empenho.save(pdf_empenho_path)

        pdf = pdfquery.PDFQuery(pdf_empenho_path)
        pdf.load()
        
        # Definir as coordenadas dos elementos desejados
        coordinatesEmpenho = [
            #nota de empenho
            {'left': 200.0, 'top': 549.52, 'width': 16.68, 'height': 10.0},    # Coordenadas do final do número de empenho
            {'left': 41.0, 'top': 418.52, 'width': 374.62, 'height': 10.0},    # Coordenadas do fornecedor (nome e CNPJ) da nota de empenho
            {'left': 421.0, 'top': 642.52, 'width': 50.02, 'height': 10.0},    # Coordenadas do valor da nota de empenho
            {'left': 200.0, 'top': 464.52, 'width': 139.57, 'height': 10.0},   #Coordenadas do número do processo
            {'left': 200.0, 'top': 503.52, 'width': 56.7, 'height': 10.0},     # Coordenadas da fonte de despesa
            {'left': 43.0, 'top': 627.52, 'width': 387.29, 'height': 10.0},    # Coordenadas da natureza da despesa
            {'left': 125.0, 'top': 306.52, 'width': 122.66, 'height': 10.0} ,   # Modalidade da licitação
            {'left': 122.0, 'top': 503.52, 'width': 33.36, 'height': 10.0} ,   # Coordenadas do PTRES        
            {'left': 296.0, 'top': 503.52, 'width': 33.36, 'height': 10.0} ,   # Coordenadas do nº da natureza da despesa
            {'left': 485.0, 'top': 503.52, 'width': 73.88, 'height': 10.0} ,   # Coordenadas do plano interno
            {'left': 407.0, 'top': 503.52,'width': 33.36, 'height': 10.0} ,   # UGR 
        ]

        
        dados_empenho = []
        for j, coord in enumerate(coordinatesEmpenho):
            element = pdf.pq(
                f'LTTextLineHorizontal:in_bbox("{coord["left"]}, {coord["top"]}, {coord["left"]+coord["width"]}, {coord["top"]+coord["height"]}")'
            )
            text = element.text().strip()
            dados_empenho.append(text)
            # Certifique-se de escrever na célula "principal" de qualquer mesclagem
            cell = ws.cell(row=2, column=j + 1)
            if cell.coordinate in ws.merged_cells:
                # pega a primeira célula do intervalo mesclado
                merged_range = [rng for rng in ws.merged_cells.ranges if cell.coordinate in rng][0]
                first_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                first_cell.value = text
            else:
                cell.value = text

        empenho_number = dados_empenho[0] if dados_empenho else "SemEmpenho"
        
        coordinatesBoleto = [
            #boleto 
            {'left': 131.0, 'top':348.99,'width': 66.164, 'height': 7.0},  #Nosso número 
            {'left': 343.0, 'top': 348.99,'width': 32.284, 'height': 7.0},     # valor boleto
            {'left': 11.0, 'top': 133.461,'width': 83.664, 'height': 7.0},     # processo BB
            {'left': 11.0, 'top': 133.461,'width': 144.333, 'height': 7.0},     # serviço BB
            {'left': 15.0, 'top': 804.344,'width': 247.688, 'height': 8.0}     # boleto BB
        ]

        start_row = 5
        for i, boleto in enumerate(boletos):
            boleto_path = os.path.join(pdfs_directory, boleto.filename)
            boleto.save(boleto_path)

            pdf_boleto = pdfquery.PDFQuery(boleto_path)
            pdf_boleto.load()

            for j, coord in enumerate(coordinatesBoleto):
                element = pdf_boleto.pq(
                    f'LTTextLineHorizontal:in_bbox("{coord["left"]}, {coord["top"]}, {coord["left"]+coord["width"]}, {coord["top"]+coord["height"]}")'
                )
                text = element.text().strip()
                ws.cell(row=start_row + i, column=j + 1).value = text

            # ✅ Inserir índice + nome do arquivo nas colunas seguintes
            last_col = len(coordinatesBoleto)  # Última coluna com dados do boleto
            index_col = last_col + 1           # Coluna para índice
            filename_col = last_col + 2        # Coluna para nome do arquivo

            ws.cell(row=start_row + i, column=index_col).value = i + 1
            ws.cell(row=start_row + i, column=filename_col).value = boleto.filename

        
        # --- 5. Salvar Excel final ---
        new_filename = f"Rascunho-{empenho_number}.xlsx"
        final_path = os.path.join(drafts_directory, new_filename)
        wb.save(final_path)
        wb.close()

        return redirect(url_for('download_excel', filename=new_filename))

    return render_template('index.html')

@app.route('/download_excel/<filename>', methods=['GET'])
def download_excel(filename):
    drafts_directory = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'RascunhosGerados')
    copied_path = os.path.join(drafts_directory, filename)
    return send_file(copied_path, as_attachment=True)

if __name__ == '__main__':
    directory_path = os.path.dirname(os.path.abspath(__file__))
    pdfs_directory = os.path.join(directory_path, 'pdfs')
    drafts_directory = os.path.join(directory_path, 'RascunhosGerados')

    # Esvaziar a pasta pdfs e a pasta RascunhosGerados antes de iniciar o servidor
    empty_folder(pdfs_directory)
    empty_folder(drafts_directory)
    
    webbrowser.open('http://localhost:5000') 
    print('Servidor iniciado na porta http://localhost:5000 !') 
    print('App desenvolvido por: RENATA VERAS VENTURIM') 

    flask_thread = threading.Thread(target=app.run)
    flask_thread.start()
    flask_thread.join()


